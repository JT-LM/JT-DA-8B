import pandas as pd
import openpyxl
from typing import List, Dict

def get_merged_cell_value(sheet, row, column):
    cell = sheet.cell(row=row, column=column)
    if cell.value is None or str(cell.value).strip() == '':
        for merged_range in sheet.merged_cells.ranges:
            if cell.coordinate in merged_range:
                return sheet.cell(row=merged_range.min_row, column=merged_range.min_col).value
    return cell.value

def is_part_of_merged_cell(sheet, row, column):
    cell = sheet.cell(row=row, column=column)
    for merged_range in sheet.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return merged_range.min_row
    return row

def get_header_rows(sheet) -> int:
    max_row = sheet.max_row
    max_col = sheet.max_column
    
    if max_row <= 1:
        return 1
    
    for row in range(1, max_row + 1):
        has_merged_cell = False
        for col in range(1, max_col + 1):
            cell = sheet.cell(row=row, column=col)
            for merged_range in sheet.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    has_merged_cell = True
                    break
            if has_merged_cell:
                break
        if not has_merged_cell:
            return max(1, row - 1)
    
    return 1 

def get_column_headers(sheet, max_row: int, column: int) -> List[str]:
    headers = []
    processed_rows = set() 
    
    for row in range(1, max_row + 1):
        min_row = is_part_of_merged_cell(sheet, row, column)

        if min_row in processed_rows:
            continue
            
        value = get_merged_cell_value(sheet, row, column)
        if value is not None and str(value).strip():
            headers.append(str(value).strip())
            processed_rows.add(min_row)
    
    return headers

def get_actual_columns(sheet, header_rows: int) -> int:
    max_col = sheet.max_column
    actual_cols = 0

    for col in range(1, max_col + 1):
        has_value = False
        for row in range(1, header_rows + 1):
            value = get_merged_cell_value(sheet, row, col)
            if value is not None and str(value).strip():
                has_value = True
                break
        if has_value:
            actual_cols = col
    
    return actual_cols

def process_single_sheet(sheet, file_path: str) -> pd.DataFrame:
    header_rows = get_header_rows(sheet)
    max_col = get_actual_columns(sheet, header_rows)
    
    new_columns = []
    
    for col in range(1, max_col + 1):
        column_headers = get_column_headers(sheet, header_rows, col)
        new_column = " - ".join(column_headers)
        new_columns.append(new_column)

    df = pd.read_excel(
        file_path,
        sheet_name=sheet.title,
        skiprows=header_rows-1
    )
    
    df.columns = new_columns
    
    for row in range(header_rows + 1, sheet.max_row + 1):
        for col in range(1, max_col + 1):
            cell = sheet.cell(row=row, column=col)
            value = get_merged_cell_value(sheet, row, col)
            
            if value is not None:
                df.iloc[row - header_rows - 1, col - 1] = value
    
    return df

def flatten_excel_headers(file_path: str) -> Dict[str, pd.DataFrame]:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    results = {}
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        results[sheet_name] = process_single_sheet(sheet, file_path)
    
    wb.close()
    
    return results

def process_excel_file(input_path: str):
    results = flatten_excel_headers(input_path)
    return results